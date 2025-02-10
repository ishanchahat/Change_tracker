import openpyxl
from openpyxl.styles import PatternFill
import os

# File path
file_path = "tracked_excel.xlsx"
log_path = "change_log.txt"

# Load or create workbook
def load_or_create_workbook():
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.save(file_path)
    return wb

# Log changes
def log_change(sheet, cell, old_value, new_value):
    with open(log_path, "a") as log_file:
        log_file.write(f"Sheet: {sheet}, Cell: {cell}, Old: {old_value}, New: {new_value}\n")

# Highlight changed cells
def highlight_cell(ws, cell):
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws[cell].fill = fill

def track_changes():
    wb = load_or_create_workbook()
    ws = wb.active
    
    changes = []
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment and "OLD:" in cell.comment.text:
                old_value = cell.comment.text.replace("OLD: ", "").strip()
                if str(cell.value) != old_value:
                    changes.append((cell.coordinate, old_value, cell.value))
                    log_change(ws.title, cell.coordinate, old_value, cell.value)
                    highlight_cell(ws, cell.coordinate)
                    cell.comment.text = f"OLD: {cell.value}"
            else:
                cell.comment = openpyxl.comments.Comment(f"OLD: {cell.value}", "Tracker")
    
    wb.save(file_path)
    if changes:
        print("Changes tracked:", changes)
    else:
        print("No changes detected.")

if __name__ == "__main__":
    track_changes()
